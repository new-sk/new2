import os
import sys
import platform
import sqlite3
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas.errors import DatabaseError

class my95_stock:
    """
    SQLite DB에서 지정한 테이블을 읽어 엑셀로 저장하고,
    저장된 엑셀 파일을 다시 읽어 출력하는 기능을 제공하는 클래스.

    - DB 이름은 고정: 'my_stock.db'
    - 엑셀 파일 이름은 고정: 'my_stock_excel.xlsx'
    - OS에 따라 경로 설정:
        • Windows: D:\my_stock\
        • macOS : 파이썬 소스코드(.py)가 있는 디렉토리
    """

    def __init__(self):
        """
        클래스 초기화: 운영체제에 따라 DB 및 엑셀 경로를 설정하고 테이블명을 저장함.

        Parameters:
            table_name (str): 사용할 SQLite 테이블 이름
        """
        self.os_name = platform.system()
        self.conn = None
        self.db_name = 'my_stock_2.db'
        self.excel_name = 'my_stock_excel.xlsx'
        self.dfi = []
        self.dfo = []

        sql_df9i = "SELECT * FROM tb_stock_code"
        sql_df9o = """
SELECT tso.*, tsc.siCode soCode,
       cast(((100 + tso.sobRate) / 100 * tso.soAmt) as integer)  sobPrice,
       cast(((100 + tso.sosRate) / 100 * tso.soAmt) as integer)  sosPrice
FROM tb_stock_obj tso, tb_stock_code tsc 
where tso.soName = tsc.siName 
and soDelYN <> 'Y'      
"""


        if self.os_name == 'Windows':
            # 테스트 기간 동안에는 동일하게 사용 : 맥과 윈도우의 디렉토리 
            # base_dir = r'D:\my_test'   
            base_dir = os.path.dirname(os.path.abspath(__file__))
        elif self.os_name == 'Darwin':  # macOS
            base_dir = os.path.dirname(os.path.abspath(__file__))
        else:
            print(f"[에러] 지원되지 않는 운영체제입니다: {self.os_name}")
            sys.exit(1)

        self.db_path = os.path.join(base_dir, self.db_name)
        self.excel_path = os.path.join(base_dir, self.excel_name)


        """
        SQLite DB에 연결, 연결 안되면 오류 출력 후 프로그램 종료.
        """
        if not os.path.isfile(self.db_path):
            print(f"[에러] DB 파일이 존재하지 않습니다: {self.db_path}")
            sys.exit(1)

        self.conn = sqlite3.connect(self.db_path)

    
    def save_table_to_excel(self, sheet_name=None):
        """
        지정된 SQLite 테이블을 엑셀 파일로 저장하되,
        기존 시트가 존재하면 서식은 유지하고 데이터만 덮어씀.
        - 테이블이 없거나 엑셀이 열려있을 경우도 예외 처리함.
        """
        if self.conn is None:
            print("[에러] DB 연결이 되어 있지 않습니다. connect_db() 먼저 호출하세요.")
            sys.exit(1)

        if sheet_name is None:
            print("[에러] 테이블명을 입력해 주세요. save_table_to_excel() 호출시 테이블명은 필수입니다.")
            sys.exit(1)

        try:
            df = pd.read_sql_query(f"SELECT * FROM {sheet_name}", self.conn)
        except (sqlite3.OperationalError, DatabaseError) as e:
            print(f"[에러] 테이블 '{sheet_name}'을(를) 불러올 수 없습니다: {e}")
            sys.exit(1)

        try:
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.delete_rows(2, ws.max_row)  # 기존 데이터 삭제 (헤더 제외)
                else:
                    ws = wb.create_sheet(title=sheet_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(self.excel_path)
            print(f"[완료] '{sheet_name}' 테이블을 '{self.excel_path}'의 '{sheet_name}' 시트에 저장했습니다.")

        except PermissionError:
            print(f"[에러] 엑셀 파일이 열려 있어 저장할 수 없습니다. '{self.excel_path}'를 닫고 다시 시도하세요.")
            sys.exit(1)
        except Exception as e:
            print(f"[에러] 엑셀 저장 중 알 수 없는 오류가 발생했습니다: {e}")
            sys.exit(1)
            
                
    def read_excel(self, sheet_name=None):
        """
        엑셀 파일에서 지정한 시트를 읽고 상위 5개 행을 출력함.

        Parameters:
            sheet_name (str): 읽을 시트명 (기본값: 테이블 이름과 동일)
        """
        if not os.path.isfile(self.excel_path):
            print(f"[에러] 엑셀 파일이 존재하지 않습니다: {self.excel_path}")
            return

        if sheet_name is None:
            print("[에러] 테이블명을 입력해 주세요. read_excel() 호출시 테이블명은 필수입니다.")
            sys.exit(1)

        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='openpyxl')
            print(f"\n[엑셀 읽기 결과 - '{sheet_name}' 시트 미리보기]")
            print(df.head())
        except Exception as e:
            print(f"[에러] 엑셀 파일 읽기 실패: {e}")




if __name__ == '__main__':
    stock = my95_stock()
    stock.connect_db()
    stock.save_table_to_excel('tb_stock_code') 
    stock.read_excel('tb_stock_code')           
    stock.save_table_to_excel('tb_stock_obj') 
    stock.read_excel('tb_stock_obj')           
